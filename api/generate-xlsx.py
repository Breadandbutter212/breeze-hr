from http.server import BaseHTTPRequestHandler
import json
import io
import re


def _is_formula(v):
    return isinstance(v, str) and v.strip().startswith('=')


def _clean_num(s):
    return re.sub(r'[£$€,\s]', '', s)


def _col_type(rows, ci):
    """Infer a column's type from its body cells: text / number / currency / percent."""
    cur = pct = any_ = False
    num = True
    for ri in range(1, len(rows)):
        raw = '' if ci >= len(rows[ri]) or rows[ri][ci] is None else str(rows[ri][ci]).strip()
        if not raw or _is_formula(raw):
            continue
        any_ = True
        if re.search(r'[£$€]', raw):
            cur = True
        if raw.endswith('%'):
            pct = True
        if not re.match(r'^-?[\d,]+(\.\d+)?$', re.sub(r'[£$€%\s]', '', raw)):
            num = False
    if not any_:
        return 'text'
    if pct and num:
        return 'percent'
    if cur and num:
        return 'currency'
    if num:
        return 'number'
    return 'text'


def build_workbook(title, sheets):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
    header_fill = PatternFill('solid', fgColor='2A5080')
    thin = Side(style='thin', color='D7DEEA')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    left_center = Alignment(horizontal='left', vertical='center')

    used_names = set()

    def safe_name(name, idx):
        n = re.sub(r'[\\/?*\[\]:]', ' ', str(name or ('Sheet%d' % (idx + 1)))).strip()[:31] or ('Sheet%d' % (idx + 1))
        base, k = n[:28], 2
        while n.lower() in used_names:
            n = ('%s %d' % (base, k))[:31]
            k += 1
        used_names.add(n.lower())
        return n

    if not sheets:
        sheets = [{'name': title or 'Sheet1', 'rows': [['(empty)']], 'chart': None}]
    if len(sheets) == 1 and (not sheets[0].get('name') or sheets[0].get('name') == 'Sheet1'):
        sheets[0]['name'] = title or 'Sheet1'

    for idx, s in enumerate(sheets):
        rows = [r if isinstance(r, list) else [r] for r in (s.get('rows') or [])]
        if not rows:
            rows = [['']]
        ws = wb.create_sheet(safe_name(s.get('name'), idx))
        max_cols = max(len(r) for r in rows)
        types = [_col_type(rows, c) for c in range(max_cols)]

        for ri, r in enumerate(rows):
            for ci in range(max_cols):
                raw = '' if ci >= len(r) or r[ci] is None else str(r[ci])
                cell = ws.cell(row=ri + 1, column=ci + 1)
                if ri == 0:
                    cell.value = raw
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = left_center
                    cell.border = border
                    continue
                cell.border = border
                t = types[ci]
                if _is_formula(raw):
                    cell.value = raw
                    if t == 'currency':
                        cell.number_format = u'"£"#,##0.00'
                    elif t == 'percent':
                        cell.number_format = '0.0%'
                    elif t == 'number':
                        cell.number_format = '#,##0'
                elif raw == '':
                    cell.value = None
                elif t == 'currency':
                    try:
                        cell.value = float(_clean_num(raw))
                    except ValueError:
                        cell.value = raw
                    cell.number_format = u'"£"#,##0.00'
                elif t == 'percent':
                    try:
                        cell.value = float(re.sub(r'[%\s,]', '', raw)) / 100.0
                    except ValueError:
                        cell.value = raw
                    cell.number_format = '0.0%'
                elif t == 'number':
                    try:
                        v = _clean_num(raw)
                        cell.value = int(v) if re.match(r'^-?\d+$', v) else float(v)
                    except ValueError:
                        cell.value = raw
                    cell.number_format = '#,##0'
                else:
                    cell.value = raw

        # Column widths from longest cell
        for ci in range(max_cols):
            longest = 9
            for r in rows:
                v = '' if ci >= len(r) or r[ci] is None else str(r[ci])
                longest = max(longest, len(v) + 2)
            ws.column_dimensions[get_column_letter(ci + 1)].width = min(max(longest, 9), 60)

        ws.freeze_panes = 'A2'
        if len(rows) > 1:
            ws.auto_filter.ref = 'A1:%s%d' % (get_column_letter(max_cols), len(rows))

        # Chart
        chart_type = (s.get('chart') or '').lower()
        if chart_type and len(rows) >= 3:
            num_cols = [c for c in range(1, max_cols) if types[c] in ('number', 'currency', 'percent')]
            if num_cols:
                chart = None
                if chart_type == 'pie':
                    chart = PieChart()
                    data = Reference(ws, min_col=num_cols[0] + 1, min_row=1, max_row=len(rows))
                    cats = Reference(ws, min_col=1, min_row=2, max_row=len(rows))
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(cats)
                elif chart_type == 'line':
                    chart = LineChart()
                else:
                    chart = BarChart()
                    chart.type = 'bar' if chart_type == 'bar' else 'col'
                if chart_type in ('column', 'bar', 'line'):
                    last = num_cols[-1] + 1
                    data = Reference(ws, min_col=num_cols[0] + 1, max_col=last, min_row=1, max_row=len(rows))
                    cats = Reference(ws, min_col=1, min_row=2, max_row=len(rows))
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(cats)
                if chart is not None:
                    chart.title = ws.title
                    chart.height = 8
                    chart.width = 15
                    ws.add_chart(chart, '%s2' % get_column_letter(max_cols + 2))

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


class handler(BaseHTTPRequestHandler):
    def log_message(self, format, *args):
        pass

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type, Authorization')
        self.end_headers()

    def do_POST(self):
        try:
            content_length = int(self.headers.get('Content-Length', 0))
            data = json.loads(self.rfile.read(content_length))
            title = data.get('title', 'Spreadsheet')
            sheets = data.get('sheets', [])
            if not sheets:
                self._error(400, 'No sheets provided')
                return
            xlsx_bytes = build_workbook(title, sheets)
            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', 'attachment; filename="spreadsheet.xlsx"')
            self.send_header('Content-Length', str(len(xlsx_bytes)))
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(xlsx_bytes)
        except Exception as e:
            self._error(500, str(e))

    def _error(self, code, msg):
        body = json.dumps({'error': msg}).encode()
        self.send_response(code)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Content-Length', str(len(body)))
        self.end_headers()
        self.wfile.write(body)
